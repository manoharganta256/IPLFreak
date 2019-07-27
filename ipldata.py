import os
import django
import click
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "IPL.settings")
django.setup()


def convert_excel(source, sheet_name):
    workbook = load_workbook(source)
    print(workbook.sheetnames)
    sheet = workbook.get_sheet_by_name(sheet_name)

    data = []
    for row in sheet.rows:
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        data.append(row_data)

    return data


def dump_to_matches(data):
    from iplfreak.models import Match

    data = data[1:]
    for row in data:
        print(row)
        match = Match(match_id=int(row[0]))

        match.season = int(row[1])
        match.city = row[2]
        # todo: date
        # match.date = row[3]
        match.team1 = row[4]
        match.team2 = row[5]
        match.toss_winner = row[6]
        match.toss_decision = 1 if row[7] == 'bat' else 2
        match.result = row[8]
        match.dl_applied = True if row[9] == '1' else False
        match.winner = row[10]
        match.win_by_runs = int(row[11])
        match.win_by_wickets = int(row[12])
        match.player_of_the_match = row[13]
        match.venue = row[14]
        match.umpire1 = row[15]
        match.umpire2 = row[16]
        match.save()

        del match


def dump_to_deliveries(data):
    from iplfreak.models import Deliveries, Match

    data = data[1:]
    for row_values in data:
        d = Deliveries()
        try:
            print(row_values[0])
            d.match_id = Match.objects.get(match_id=row_values[0])
            d.innings = row_values[1]
            d.batting_team = row_values[2]
            d.bowling_team = row_values[3]
            d.over = row_values[4]
            d.ball = row_values[5]
            d.batsman = row_values[6]
            d.non_striker = row_values[7]
            d.bowler = row_values[8]
            d.is_super_over = row_values[9]
            d.wide_runs = row_values[10]
            d.bye_runs = row_values[11]
            d.legbye_runs = row_values[12]
            d.noball_runs = row_values[13]
            d.penalty_runs = row_values[14]
            d.batsman_runs = row_values[15]
            d.extra_runs = row_values[16]
            d.total_runs = row_values[17]
            d.player_dismissed = row_values[18]
            d.dismissal_kind = row_values[19]
            d.fielder = row_values[20]
            d.save()
            del d
        except Exception as e:
            del d
            print(e)


@click.command()
@click.option('-m', '--matches', is_flag=True)
@click.option('-d', '--deliveries', is_flag=True)
@click.argument('source', nargs=1)
def ipl_data(matches, deliveries, source):
    print(source)
    if matches:
        data = convert_excel(source, 'matches')
        dump_to_matches(data)
    else:
        data = convert_excel(source, 'deliveries')
        dump_to_deliveries(data)


if __name__ == '__main__':
    ipl_data()